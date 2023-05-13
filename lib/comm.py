'''
共用函数
'''
import os
import time
from datetime import datetime
from functools import wraps
import ipaddress
from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl import load_workbook

# 项目根目录
BASE_PATH = os.path.dirname(os.path.dirname(__file__))

# 文件输出目录
EXPORT_PATH = os.path.join(BASE_PATH, 'EXPORT')

# 定义目录名称为当天日期（格式：20220609）
dir_name = datetime.now().strftime("%Y%m%d")

# 目录创建
new_path = os.path.join(EXPORT_PATH, dir_name)
if not os.path.isdir(new_path):
    os.makedirs(new_path)
backup_path = '%s\\%s\\export_conf' % (EXPORT_PATH, dir_name)
config_path = '%s\\%s\\modify_conf' % (EXPORT_PATH, dir_name)
generate_table = '%s\\%s\\generate_table' % (EXPORT_PATH, dir_name)
diff_config_path = '%s\\%s\\diff_conf' % (EXPORT_PATH, dir_name)
if not os.path.isdir(backup_path):
    os.makedirs(backup_path)
if not os.path.isdir(config_path):
    os.makedirs(config_path)
if not os.path.isdir(generate_table):
    os.makedirs(generate_table)
if not os.path.isdir(diff_config_path):
    os.makedirs(diff_config_path)


# 记录程序执行时间装饰器
def timer(func):

    @wraps(func)
    def wrapper(*args, **kwargs):
        start_time = datetime.now()
        func(*args, **kwargs)
        end_time = datetime.now()
        # print('\n' + '-' * 42)
        print('\n执行完毕，共耗时 {:0.2f} 秒.'.format(
            (end_time - start_time).total_seconds()))
        # print('+' * 80)
        # return res

    return wrapper


# 记录运行结果的装饰器
def result_count(func):

    @wraps(func)
    def wrapper(*args, **kwargs):

        hosts, failed_hosts, task_desc = func(*args, **kwargs)
        print('-' * 42)
        print('\n设备总数 {} 台，成功 {} 台，失败 {} 台.'.format(
            len(hosts),
            len(hosts) - len(failed_hosts), len(failed_hosts)))
        print(
            f'\nFailed_hosts list see in : \"{EXPORT_PATH}\\{dir_name}\\result_{dir_name}.log\"\n\nLogfile see in : \"{BASE_PATH}\\nornir.log\"'
        )

        return hosts, failed_hosts, task_desc

    return wrapper


# 保存运行结果记录的装饰器
def result_write(func):

    @wraps(func)
    def wrapper(*args, **kwargs):

        hosts, failed_hosts, task_desc = func(*args, **kwargs)
        result_count = ('设备总数 {} 台，成功 {} 台，失败 {} 台.'.format(
            len(hosts),
            len(hosts) - len(failed_hosts), len(failed_hosts)))

        # 计算执行成功的设备列表
        success_hosts = list(set(hosts) - set(failed_hosts))

        global time_str
        # time_str = datetime.now()
        time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
        with open(os.path.join(EXPORT_PATH + '\\' + dir_name + '\\',
                               f'result_{dir_name}.log'),
                  'a',
                  encoding="utf-8") as f:
            log_title = task_desc.center(100, '=') + '\n' + time_str.center(
                100, '=') + '\n'
            f.write(log_title)
            f.write(result_count + '\n')
            f.write('\n执行成功设备列表：\n')
            for i in success_hosts:
                f.write(f'{i[0]}, {i[1]}\n')

            f.write('\nNG设备列表：\n')
            for i in failed_hosts:
                f.write(f'{i[0]}, {i[1]}\n')
                # f.write('\n')
            f.write('\n')

        return hosts, failed_hosts, task_desc

    return wrapper


# 生成用于统计主机的列表
def create_count_list(nr_obj, failed_hosts):
    """
    根据inventory nr对象和执行task后返回的Result（failed_hosts），
    返回两个列表：
     - hosts_list
     - failed_hosts_list
    数据格式为列表套元组：
    - list = [(name, hostname), (name, hostname), ...]
    """

    # 定义列表，装载（name, hostname）
    hosts_list = []
    failed_hosts_list = []

    for n, h in nr_obj.inventory.hosts.items():
        hosts_list.append((n, h.hostname))
        for i in failed_hosts:
            if i == n:
                failed_hosts_list.append((n, h.hostname))

    return hosts_list, failed_hosts_list


# 判断是否是正确格式的IP地址，IP地址网络，IP地址范围
def is_valid_ipv4_input(ipv4_str):

    try:
        # 尝试将输入解析为 IPv4Address
        ipaddress.IPv4Address(ipv4_str)
        return True
    except ValueError:
        return False
        # try:
        #     # 将输入拆分为两个 IP 地址
        #     start, end = ipv4_str.split('-')
        #     # 尝试将输入解析为 IPv4Address
        #     ipaddress.IPv4Address(start.strip())
        #     ipaddress.IPv4Address(end.strip())
        #     return True
        # except ValueError as e:
        #     # 尝试将输入解析为 IPv4Network
        #     try:
        #         network = ipaddress.IPv4Network(ipv4_str)
        #         if network.hostmask != '0.0.0.0':
        #             return True
        #         else:
        #             return False
        #     except ValueError as e:
        #         # print(e)
        #         return False


# 合并第一列相同内容的单元格，居中
def excel_style(file_path):
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    # 遍历第一列并建立字典，键为值，值为包含该值所有单元格的列表
    merge_dict = {}

    for i, row in enumerate(
            worksheet.iter_rows(min_row=2, min_col=1, values_only=True)):
        key = row[0]
        if key not in merge_dict:
            merge_dict[key] = [i + 2]
        else:
            merge_dict[key].append(i + 2)

    # 遍历字典，将相同内容的单元格进行合并
    for rows in merge_dict.values():
        if len(rows) > 1:
            # 合并单元格
            first_index = rows[0]
            last_index = rows[-1]
            worksheet.merge_cells(start_row=first_index,
                                  end_row=last_index,
                                  start_column=1,
                                  end_column=1)

    # 修改列宽
    worksheet.column_dimensions['A'].width = 50
    worksheet.column_dimensions['B'].width = 16
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['D'].width = 10
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 10
    # 单元格垂直居中
    for row in worksheet:
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                       vertical='center')

    # 将所有表头单元格的值转换为大写
    for cell in worksheet[1]:
        cell.value = str(cell.value).upper()

    # 迭代每个非空列并调整列宽
    for column in worksheet.columns:
        # 获取当前列名称
        column_name = column[0].column_letter

        # 计算此列中所有单元格中包含的最长文本
        max_length = 0
        for cell in column:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except TypeError:
                pass

        # 根据最长文本设置列宽（添加 2 作为 buffer）
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_name].width = adjusted_width

    # 保存
    workbook.save(file_path)


# 处理DataFrame，根据results聚合结果中的result内容 组合成一整个DataFrame 写入表格
def concat_dataframe(results, file_path):
    # 定义一个列表装 DataFrame，每个DataFrame是一台设备上的mac地址表
    df_list = []
    try:
        for i in results.keys():
            df = results[i][0].result
            # 判断result是否字符串，如是则表示是failed_hosts的设备
            if not isinstance(df, str):
                # 将单个DataFrame追加到列表 df_list
                df_list.append(df)

        # 将多个DataFrame合并，每个DataFrame是一台设备的mac地址表
        df_all = pd.concat(df_list)
        # 写入表格
        df_all.to_excel(file_path, index=False)

        # 表格样式修改
        excel_style(file_path)

    except Exception as e:
        print(f'有错误：{e}')


# 对第一列中已合并的单元格拆分，拆分后的单元格数据与拆分前相同
def split_cells(file_path):

    # 读取 Excel 表格
    df = pd.read_excel(file_path)

    # 获取第一列的列名
    col_name = df.columns[0]

    # 拆分第一列的合并单元格
    df[col_name] = df[col_name].ffill()

    # 返回DataFrame
    return df
